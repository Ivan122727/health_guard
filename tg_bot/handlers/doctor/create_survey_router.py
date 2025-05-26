import os
import uuid
from aiogram import F, Router
from aiogram.types import CallbackQuery, Message, FSInputFile
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from shared.sqlalchemy_db_.sqlalchemy_model.user import UserDBM
from tg_bot.handlers.common.message_service import MessageService
from tg_bot.handlers.doctor.create_survey_service import CreateSurveyService
from tg_bot.keyboards import DoctorAction, DoctorKeyboard
from tg_bot.blanks import DoctorBlank
from tg_bot.states.survey import CreateSurveyStates

router = Router()

@router.callback_query(F.data.startswith(DoctorAction.CREATE_TITLE_SURVEY.value))
async def handle_create_title_survey(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await callback_query.answer()
    
    count_questions = await CreateSurveyService.get_count_questions_in_survey(state)
    
    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=callback_query.from_user.id,
        text=blank.get_survey_title_input_blank(),
        reply_markup=keyboard.get_question_management_keyboard(count_questions),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=callback_query.message,
        new_state=CreateSurveyStates.waiting_new_title_survey
    )

async def proccess_edit_survey_title(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    change_title: bool = True
):
    if change_title:
        title = message.text
    else:
        title = await CreateSurveyService.get_confirmed_survey_title(state)

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_survey_title_confirmation_blank(title),
        reply_markup=keyboard.get_survey_title_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        new_state=CreateSurveyStates.waiting_new_title_survey
    )

    await CreateSurveyService.save_not_confirmed_title(
        state=state,
        title=title,
    )

    if change_title:
        await MessageService.remove_previous_message(
            message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

@router.message(CreateSurveyStates.waiting_new_title_survey)
async def handle_input_title_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_edit_survey_title(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        change_title=True
    )

@router.callback_query(F.data.startswith(DoctorAction.EDIT_SURVEY_TITLE.value))
async def handle_edit_survey_title(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_edit_survey_title(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        change_title=False,
    )

async def proccess_handle_choose_type_question(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True
):
    count_questions = await CreateSurveyService.get_count_questions_in_survey(state)

    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_choose_type_survey_blank(),
        reply_markup=keyboard.get_question_type_selection_keyboard(count_questions),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=CreateSurveyStates.waiting_choose_type_question
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )    

@router.callback_query(F.data.startswith(DoctorAction.CONFIRM_TITLE_SURVEY.value))
async def handle_confirm_title_survey(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await CreateSurveyService.confirm_survey_title(
        state=state
    )

    await proccess_handle_choose_type_question(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm
    )

@router.callback_query(F.data.startswith(DoctorAction.CHOOSE_TYPE_QUESTION.value))
async def handle_choose_type_question_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_handle_choose_type_question(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm
    )

@router.message(CreateSurveyStates.waiting_choose_type_question)
async def handle_choose_type_question(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_handle_choose_type_question(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False,
    )

async def proccess_choose_new_question_select(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True
):
    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_create_from_scratch_blank(),
        reply_markup=keyboard.get_confirm_create_new_question_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=CreateSurveyStates.waiting_create_new_question_select,
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id
        )

@router.callback_query(F.data.startswith(DoctorAction.CREATE_NEW_QUESTION.value))
async def handle_choose_new_question_type_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_choose_new_question_select(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm
    )

@router.message(CreateSurveyStates.waiting_create_new_question_select)
async def handle_choose_new_question_type(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_choose_new_question_select(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False
    )

async def proccess_choose_template_question_select(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True
):
    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_use_template_blank(),
        reply_markup=keyboard.get_confirm_create_template_question_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=CreateSurveyStates.waiting_create_template_question_select
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id
        )

@router.callback_query(F.data.startswith(DoctorAction.CREATE_TEMPLATE_QUESTION.value))
async def handle_choose_template_question_type_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_choose_template_question_select(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(CreateSurveyStates.waiting_create_template_question_select)
async def handle_choose_template_question_type(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_choose_template_question_select(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False,
    )

@router.callback_query(F.data.startswith(DoctorAction.CONFIRM_CREATE_NEW_QUESTION.value))
async def handle_confirm_new_question_type(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    count_questions = await CreateSurveyService.get_count_questions_in_survey(state)

    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=callback_query.from_user.id,
        text=blank.get_create_from_scratch_blank(step=blank.STATE_QUESTION_TEXT),
        reply_markup=keyboard.get_question_management_keyboard(count_questions=count_questions),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=callback_query.message,
        new_state=CreateSurveyStates.waiting_new_question_text,
    )

@router.callback_query(F.data.startswith(DoctorAction.CONFIRM_CREATE_TEMPLATE_QUESTION.value))
async def handle_confirm_template_question_type(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    count_questions = await CreateSurveyService.get_count_questions_in_survey(state)
    
    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=callback_query.from_user.id,
        text=blank.get_use_template_blank(),
        reply_markup=keyboard.get_question_management_keyboard(count_questions),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=callback_query.message,
        new_state=CreateSurveyStates.waiting_template_question_id
    )

# Добавление текста вопроса в новый опрос
@router.message(CreateSurveyStates.waiting_new_question_text)
async def handle_add_new_question_text(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    # Добавляем текст в опрос
    await CreateSurveyService.add_or_edit_question_text_in_survey(
        state=state,
        text=message.text,
    )

    count_questions = await CreateSurveyService.get_count_questions_in_survey(state)

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=message.from_user.id,
        text=blank.get_create_from_scratch_blank(step=blank.STATE_QUESTION_OPTIONS, current_question=message.text),
        reply_markup=keyboard.get_question_management_keyboard(count_questions),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        new_state=CreateSurveyStates.waiting_new_question_options
    )
    
    # Удаление сообщения пользователя
    await MessageService.remove_previous_message(
        bot=message.bot, 
        user_id=message.from_user.id,
        message_id=message.message_id
    )

# Добавление вариантов ответа на вопрос в новый опрос
@router.message(CreateSurveyStates.waiting_new_question_options)
async def handle_add_new_question_options(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    count_questions = await CreateSurveyService.get_count_questions_in_survey(state=state)
    
    current_question = await CreateSurveyService.get_current_question(state)

    if await CreateSurveyService.is_valid_answer_options(message.text):
        answer_options = await CreateSurveyService.parse_answer_options(
            raw_options=message.text
        )
        await CreateSurveyService.add_options_to_current_question(
            state=state,
            answer_options=answer_options
        )
        
        count_questions = await CreateSurveyService.get_count_questions_in_survey(state=state)

        text=blank.get_create_from_scratch_blank(step=blank.STATE_QUESTION_TEXT, count_questions=count_questions)
        new_state = CreateSurveyStates.waiting_new_question_text
    else:
        text=blank.get_create_from_scratch_blank(step=blank.STATE_QUESTION_OPTIONS, current_question=current_question.text)
        new_state=CreateSurveyStates.waiting_new_question_options

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=message.from_user.id,
        text=text,
        reply_markup=keyboard.get_question_management_keyboard(count_questions),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        new_state=new_state
    )
    
    # Удаление сообщения пользователя
    await MessageService.remove_previous_message(
        bot=message.bot, 
        user_id=message.from_user.id,
        message_id=message.message_id
    )

@router.message(CreateSurveyStates.waiting_template_question_id)
async def handle_add_template_question(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    template_question_id = await MessageService.get_value_from_callback_data(
        message.text,
        position_index=1,
        default_value=-1,
        type=int,
    )
    
    is_success = await CreateSurveyService.add_or_edit_template_question_to_survey(
        state=state,
        user_id=message.from_user.id,
        template_question_id=template_question_id
    )
    
    count_questions = await CreateSurveyService.get_count_questions_in_survey(state)

    if is_success:
        await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=message.from_user.id,
            text=blank.get_use_template_blank(),
            reply_markup=keyboard.get_question_management_keyboard(count_questions),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=CreateSurveyStates.waiting_template_question_id,
        )
    else:
        await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=message.from_user.id,
            text=blank.get_use_template_blank(is_error=True),
            reply_markup=keyboard.get_question_management_keyboard(count_questions),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=CreateSurveyStates.waiting_template_question_id,
        )
    
    # Удаление сообщения пользователя
    await MessageService.remove_previous_message(
        bot=message.bot, 
        user_id=message.from_user.id,
        message_id=message.message_id
    )

@router.callback_query(F.data.startswith(DoctorAction.SAVE_SURVEY.value))
async def handle_save_survey(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    survey_was_added = await CreateSurveyService.save_survey_in_db(
        state=state,
        user_id=callback_query.from_user.id,
    )

    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=callback_query.from_user.id,
        text=blank.get_default_blank(user_dbm.full_name),
        reply_markup=keyboard.get_default_keyboard(),
        state=state,
        previous_message_key="start_message_id",
        message_id_storage_key="start_message_id",
        message=callback_query.message,
        new_state=None
    )    

    if survey_was_added:
        await callback_query.answer("Опрос был создан!")
    else:
        await callback_query.answer("Создание опроса было отменено!")

@router.callback_query(F.data.startswith(DoctorAction.CANCEL_CREATE_SURVEY.value))
async def handle_cancel_create_survey(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=callback_query.from_user.id,
        text=blank.get_default_blank(user_dbm.full_name),
        reply_markup=keyboard.get_default_keyboard(),
        state=state,
        previous_message_key="start_message_id",
        message_id_storage_key="start_message_id",
        message=callback_query.message,
        new_state=None
    )    

    await CreateSurveyService.clear_survey_data(
        state=state,
        user_id=callback_query.from_user.id,
    ) 

    await callback_query.answer("Создание опроса было отменено!")


async def proccess_handle_edit_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True
):
    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_question_info(
            current_question=await CreateSurveyService.get_current_question(state),
            question_number=await CreateSurveyService.get_current_question_number(state),
            total_questions=await CreateSurveyService.get_count_questions_in_survey(state),
        ),
        reply_markup=keyboard.get_edit_survey_keyboard(
            current_question=await CreateSurveyService.get_current_question(state),
            previout_question=await CreateSurveyService.get_previous_question(state),
            next_question=await CreateSurveyService.get_next_question(state)
        ),
        state=state,
        previous_message_key="start_message_id",
        message_id_storage_key="start_message_id",
        message=message_from_cq,
        new_state=CreateSurveyStates.edit_survey
    )

    if from_cq == False:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

@router.callback_query(F.data.startswith(DoctorAction.EDIT_SURVEY.value))
async def handle_edit_survey_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_handle_edit_survey(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(CreateSurveyStates.edit_survey)
async def handle_edit_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_handle_edit_survey(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False,
    )

@router.callback_query(F.data.startswith(DoctorAction.SET_CURRENT_QUESTION.value))
async def handle_set_current_question(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    question_id = await MessageService.get_value_from_callback_data(
        callback_data=callback_query.data,
        default_value=None,
        type=str,
    )
    
    if await CreateSurveyService.get_question_by_id(state, question_id):
        await CreateSurveyService.set_current_question(
            state=state,
            question_id=question_id
        )
    
    await proccess_handle_edit_survey(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.callback_query(F.data.startswith(DoctorAction.REMOVE_CURRENT_QUESTION.value))
async def handle_remove_current_question(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await CreateSurveyService.remove_current_question(state)
    
    await proccess_handle_edit_survey(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

async def process_handle_create_question(
        callback_query: CallbackQuery,
        state: FSMContext,
        keyboard: type[DoctorKeyboard],
        blank: type[DoctorBlank],
        user_dbm: type[UserDBM],
        save_edit_question: bool = False,
):
    current_question = await CreateSurveyService.get_current_question(state)

    count_questions = await CreateSurveyService.get_count_questions_in_survey(state)

    if save_edit_question:
        await CreateSurveyService.save_edit_question_id(
            state=state,
            question_id=current_question.id
        )

    if current_question.is_from_template:
        text=blank.get_use_template_blank()
        new_state=CreateSurveyStates.waiting_template_question_id
    else:
        text=blank.get_create_from_scratch_blank(step=blank.STATE_QUESTION_TEXT)
        new_state=CreateSurveyStates.waiting_new_question_text

    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=callback_query.from_user.id,
        text=text,
        reply_markup=keyboard.get_question_management_keyboard(count_questions=count_questions),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=callback_query.message,
        new_state=new_state,
    )

@router.callback_query(F.data.startswith(DoctorAction.FINISH_EDITING.value))
async def handle_finish_editing(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await process_handle_create_question(
        callback_query=callback_query,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.callback_query(F.data.startswith(DoctorAction.EDIT_QUESTION.value))
async def handle_edit_current_question(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await process_handle_create_question(
        callback_query=callback_query,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        save_edit_question=True
    )

@router.callback_query(F.data.startswith(DoctorAction.GET_LIST_QUESTIONS.value))
async def handle_get_list_questions(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await callback_query.answer()
    wb = Workbook()
    ws = wb.active
    ws.title = "Список вопросов"
    
    # Стили форматирования
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), 
                  right=Side(style='thin'), 
                  top=Side(style='thin'), 
                  bottom=Side(style='thin'))
    alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
    
    # Заголовки
    headers = ["ID вопроса", "Текст вопроса", "Варианты ответов", "Создатель", "Публичный"]
    ws.append(headers)
    
    # Форматирование заголовков
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    # Получаем данные
    question_dbms = await CreateSurveyService.get_available_questions(user_dbm.tg_id)

    # Заполняем данными
    for question_dbm in question_dbms:
        answer_options = "\n".join(question_dbm.answer_options) if question_dbm.answer_options else ""
        row = [
            question_dbm.id, 
            question_dbm.question_text, 
            answer_options, 
            question_dbm.created_by, 
            "✓" if question_dbm.is_public else "✗"
        ]
        ws.append(row)
        
        # Форматирование строки данных
        current_row = ws.max_row
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.alignment = alignment
            cell.border = border
            
            # Центрируем ID и статус публичности
            if col in [1, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Автоматическая ширина столбцов с ограничением
    col_widths = {
        'A': 10,  # ID вопроса
        'B': 40,  # Текст вопроса
        'C': 30,  # Варианты ответов
        'D': 15,  # Создатель
        'E': 10   # Публичный
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Автоматическая высота строк
    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                lines = cell.value.count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
        ws.row_dimensions[row[0].row].height = 20 * max_lines
    
    # Замораживаем заголовки
    ws.freeze_panes = 'A2'
    
    # Добавляем фильтры
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    
    # Сохраняем и отправляем файл
    file_path = f'./questions_list_{uuid.uuid4()}.xlsx'
    wb.save(file_path)
    
    try:
        stat_file = FSInputFile(file_path)
        await callback_query.message.answer_document(
            document=stat_file,
            caption="Список всех доступных вопросов"
        )
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)