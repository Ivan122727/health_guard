import os
from typing import Optional
import uuid
from aiogram import F, Router
from aiogram.types import CallbackQuery, Message, FSInputFile
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from shared.sqlalchemy_db_.sqlalchemy_model import UserDBM, ScheduledSurveyDBM, SurveyDBM
from tg_bot.handlers.common.message_service import MessageService
from tg_bot.keyboards import DoctorAction, DoctorKeyboard
from tg_bot.blanks import DoctorBlank
from tg_bot.states.survey import ScheduleSurveyStates
from tg_bot.handlers.doctor.schedule_survey_service import ScheduleSurveyService

router = Router()

@router.callback_query(F.data.startswith(DoctorAction.CANCEL_SCHEDULING.value))
async def hadle_cancel_scheduling(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await ScheduleSurveyService.clear_schedule_data(
        state=state,
    )

    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_default_blank(user_dbm.full_name),
        reply_markup=keyboard.get_default_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=callback_query.message,
        new_state=None
    )

async def proccess_hadle_shoose_type_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True
):
    await ScheduleSurveyService.clear_schedule_data(
        state=state,
    )
    
    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_survey_scheduling_blank(),
        reply_markup=keyboard.get_survey_schedule_type_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=ScheduleSurveyStates.waiting_choose_type_survey
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )    

@router.callback_query(F.data.startswith(DoctorAction.SCHEDULE_SURVEY.value))
async def handle_choose_type_survey_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_hadle_shoose_type_survey(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(ScheduleSurveyStates.waiting_choose_type_survey)
async def handle_choose_type_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_hadle_shoose_type_survey(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False,
    )


async def proccess_hadle_choose_multiple_times_per_day(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True
):
    message_from_cq = message if from_cq else None

    await ScheduleSurveyService.save_frequency_type_survey(
        state=state,
        frequency_type=ScheduledSurveyDBM.FrequencyType.MULTIPLE_TIMES_PER_DAY
    )

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_choose_multiple_times_blank(),
        reply_markup=keyboard.get_multiple_times_per_day_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=ScheduleSurveyStates.waiting_choose_multiple_times_per_day
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        ) 

@router.callback_query(F.data.startswith(DoctorAction.CHOOSE_MULTIPLE_TIMES_PER_DAY.value))
async def hadle_choose_multiple_times_per_day_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_hadle_choose_multiple_times_per_day(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(ScheduleSurveyStates.waiting_choose_multiple_times_per_day)
async def hadle_choose_multiple_times_per_day(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_hadle_choose_multiple_times_per_day(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False,
    )


async def proccess_handle_set_times_per_day(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True,
    times_per_day: Optional[int] = None,
    error_msg: Optional[str] = None
):
    await ScheduleSurveyService.set_times_per_day(
        state=state,
        times_per_day=times_per_day,
    )


    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_set_times_per_day_blank(await ScheduleSurveyService.get_times_per_day(state), error_msg),
        reply_markup=keyboard.get_multiple_times_per_day_keyboard(flag=False),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=ScheduleSurveyStates.waiting_set_times_per_day
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

@router.callback_query(F.data.startswith(DoctorAction.SET_TIMES_PER_DAY.value))
async def hadle_choose_times_per_day_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    times_per_day = await MessageService.get_value_from_callback_data(
        callback_data=callback_query.data,
        default_value=None,
    ) 

    await proccess_handle_set_times_per_day(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        times_per_day=times_per_day
    )

@router.message(ScheduleSurveyStates.waiting_set_times_per_day)
async def hadle_choose_times_per_day(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    count = await ScheduleSurveyService.get_times_per_day(state=state)
    
    is_success, schedule_times, error_msg = await ScheduleSurveyService.validate_and_parse_times(message.text, count)

    if is_success:
        await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_survey_period_blank(),
            reply_markup=keyboard.get_date_period_keyboard(),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_survey_period
        )
        
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

        await ScheduleSurveyService.save_times(
            state=state,
            schedule_times=schedule_times,
        )
    else:
        await proccess_handle_set_times_per_day(
            message=message,
            state=state,
            keyboard=keyboard,
            blank=blank,
            user_dbm=user_dbm,
            from_cq=False,
            error_msg=error_msg,
        )


async def proccess_hadle_select_once_per_day(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True,
    error_msg: Optional[str] = None
):
    await ScheduleSurveyService.save_frequency_type_survey(
        state=state,
        frequency_type=ScheduledSurveyDBM.FrequencyType.ONCE_PER_DAY
    )
    
    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_once_per_day_blank(error_msg),
        reply_markup=keyboard.get_back_to_schedule_type_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=ScheduleSurveyStates.waiting_choose_once_per_day
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )    

@router.callback_query(F.data.startswith(DoctorAction.CHOOSE_ONCE_PER_DAY.value))
async def hadle_select_once_per_day_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_hadle_select_once_per_day(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(ScheduleSurveyStates.waiting_choose_once_per_day)
async def hadle_select_once_per_day(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    is_success, schedule_times, error_msg = await ScheduleSurveyService.validate_and_parse_times(message.text, 1)
    if is_success:
        await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_survey_period_blank(),
            reply_markup=keyboard.get_date_period_keyboard(),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_survey_period
        )

        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

        await ScheduleSurveyService.save_times(
            state=state,
            schedule_times=schedule_times,
        )
    else:
        await proccess_hadle_select_once_per_day(
            message=message,
            state=state,
            keyboard=keyboard,
            blank=blank,
            user_dbm=user_dbm,
            from_cq=False,
            error_msg=error_msg,
        )


async def proccess_hadle_select_every_few_days(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True
):
    await ScheduleSurveyService.save_frequency_type_survey(
        state=state,
        frequency_type=ScheduledSurveyDBM.FrequencyType.EVERY_FEW_DAYS
    )

    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_every_few_days_blank(),
        reply_markup=keyboard.get_every_few_days_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=ScheduleSurveyStates.waiting_choose_every_few_days
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )    

@router.callback_query(F.data.startswith(DoctorAction.CHOOSE_EVERY_FEW_DAYS.value))
async def hadle_select_every_few_days_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_hadle_select_every_few_days(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(ScheduleSurveyStates.waiting_choose_every_few_days)
async def hadle_select_every_few_days(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await proccess_hadle_select_every_few_days(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False,
    )


async def proccess_handle_set_interval_days(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True,
    interval_days: Optional[int] = None,
    error_msg: Optional[str] = None
):
    await ScheduleSurveyService.set_interval_days(
        state=state,
        interval_days=interval_days,
    )

    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
        bot=message.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_every_few_days_time_blank(await ScheduleSurveyService.get_interval_days(state), error_msg),
        reply_markup=keyboard.get_multiple_times_per_day_keyboard(flag=False),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=message_from_cq,
        new_state=ScheduleSurveyStates.waiting_set_interval_days
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

@router.callback_query(F.data.startswith(DoctorAction.SET_INTERVAL_DAYS.value))
async def handle_set_interval_days_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    interval_days = await MessageService.get_value_from_callback_data(
        callback_data=callback_query.data,
        default_value=None,
    ) 

    await proccess_handle_set_interval_days(
        message=callback_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        interval_days=interval_days,
    )

@router.message(ScheduleSurveyStates.waiting_set_interval_days)
async def handle_set_interval_days(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    is_success, schedule_times, error_msg = await ScheduleSurveyService.validate_and_parse_times(message.text, 1)
    
    if is_success:
        await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_survey_period_blank(),
            reply_markup=keyboard.get_date_period_keyboard(),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_survey_period
        )

        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

        await ScheduleSurveyService.save_times(
            state=state,
            schedule_times=schedule_times
        )
    else:
        await proccess_handle_set_interval_days(
            message=message,
            state=state,
            keyboard=keyboard,
            blank=blank,
            user_dbm=user_dbm,
            from_cq=False,
            error_msg=error_msg
        )


@router.message(ScheduleSurveyStates.waiting_survey_period)
async def handle_set_survey_period(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    start_date_str = await MessageService.get_value_from_callback_data(
        callback_data=message.text, 
        position_index=1, 
        default_value=None,
        sep="-",
        type=str
    )

    end_date_str = await MessageService.get_value_from_callback_data(
        callback_data=message.text, 
        position_index=2, 
        default_value=None,
        sep="-",
        type=str
    )

    is_success, start_date, end_date, error_msg =  await ScheduleSurveyService.validate_and_parse_survey_dates(
        start_date_str, 
        end_date_str
    )

    if is_success:
        has_both_dates = True
        await ScheduleSurveyService.save_survey_period(
            state=state,
            start_date=start_date,
            end_date=end_date,
        )
    else:
        has_both_dates = False

    await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_survey_period_blank(
                error_msg=error_msg,
                current_start_date=start_date,
                current_end_date=end_date,
            ),
            reply_markup=keyboard.get_date_period_keyboard(has_both_dates=has_both_dates),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_survey_period
        )

    await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
    )

async def proccess_handle_survey_selection(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    page: int = 0,
    from_cq: bool = False,
):
    survey_dbms = await ScheduleSurveyService.get_available_surveys(
        user_id=user_dbm.tg_id,
    )

    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_survey_selection_blank(len(survey_dbms)),
            reply_markup=keyboard.get_survey_selection_keyboard(
                survey_dbms=survey_dbms,
                page=page,
            ),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_select_survey,
            message=message_from_cq,
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )

@router.callback_query(F.data.startswith(DoctorAction.CONFIRM_DATE_PERIOD.value))
async def handle_survey_selection_cq(
    call_back_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    current_page = await MessageService.get_value_from_callback_data(
        callback_data=call_back_query.data,
        default_value=0,
    )

    await ScheduleSurveyService.save_select_survey_current_page(
        state=state,
        page=current_page,
    )

    await proccess_handle_survey_selection(
        message=call_back_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        page=current_page,
        from_cq=True,
    )

@router.message(ScheduleSurveyStates.waiting_select_survey)
async def handle_survey_selection(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    current_page = await ScheduleSurveyService.get_select_survey_current_page(state)

    await proccess_handle_survey_selection(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        page=current_page,
    )


async def proccess_confirm_selected_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    survey_dbm: Optional[SurveyDBM] = None,
    from_cq: bool = False
):
    message_from_cq = message if from_cq else None

    await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_survey_planning_template(survey_dbm=survey_dbm),
            reply_markup=keyboard.get_survey_confirmation_keyboard(survey_id=survey_dbm.id),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_confirm_selected_survey,
            message=message_from_cq,
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )


@router.callback_query(F.data.startswith(DoctorAction.SELECT_SURVEY.value))
async def show_selected_survey_cq(
    call_back_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    survey_id = await MessageService.get_value_from_callback_data(
        callback_data=call_back_query.data,
        default_value=0,
    )

    await ScheduleSurveyService.save_selected_survey(
        state=state,
        survey_id=survey_id,
        user_id=user_dbm.tg_id,
    )

    survey_dbm = await ScheduleSurveyService.get_selected_survey(state)

    await proccess_confirm_selected_survey(
        message=call_back_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=True,
        survey_dbm=survey_dbm,
    )


@router.message(ScheduleSurveyStates.waiting_confirm_selected_survey)
async def show_selected_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    survey_dbm = await ScheduleSurveyService.get_selected_survey(state)

    await proccess_confirm_selected_survey(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        survey_dbm=survey_dbm,
    )


async def proccess_handle_patient_selection(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True,
    page: int = 0
):
    message_from_cq = message if from_cq else None

    survey_dbm = await ScheduleSurveyService.get_selected_survey(state)
    patient_dbms = await ScheduleSurveyService.get_connected_patients(user_dbm.tg_id)

    await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_patient_selection_template(
                survey_dbm=survey_dbm,
                user_id=user_dbm.tg_id,
                has_patients=len(patient_dbms),
            ),
            reply_markup=keyboard.get_patient_selection_keyboard(
                patients_dbms=patient_dbms,
                page=page,
            ),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_select_patient,
            message=message_from_cq,
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )  


@router.callback_query(F.data.startswith(DoctorAction.CONFIRM_SURVEY_SELECTION.value))
async def handle_patient_selection_cq(
    call_back_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    current_page = await MessageService.get_value_from_callback_data(
        callback_data=call_back_query.data,
        default_value=0,
    )

    await ScheduleSurveyService.save_select_patient_current_page(
        state=state,
        page=current_page,
    )

    await proccess_handle_patient_selection(
        message=call_back_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        page=current_page,
    )

@router.message(ScheduleSurveyStates.waiting_select_patient)
async def handle_patient_selection(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    current_page = await ScheduleSurveyService.get_select_patient_current_page(state)

    await proccess_handle_patient_selection(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        page=current_page,
        from_cq=False,
    )


async def proccess_show_selected_patient(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True,
):
    message_from_cq = message if from_cq else None

    survey_dbm = await ScheduleSurveyService.get_selected_survey(state)
    patient_dbm = await ScheduleSurveyService.get_selected_patient(state)

    try:
        await MessageService.edith_managed_message(
                bot=message.bot,
                user_id=user_dbm.tg_id,
                text=blank.get_patient_confirmation_template(
                    survey_dbm=survey_dbm,
                    patient_dbm=patient_dbm,
                    doctor_id=user_dbm.tg_id,
                ),
                reply_markup=keyboard.get_patient_confirmation_keyboard(
                    patient_dbm=patient_dbm,
                ),
                state=state,
                previous_message_key="start_msg_id",
                message_id_storage_key="start_msg_id",
                new_state=ScheduleSurveyStates.waiting_confirm_selected_patient,
                message=message_from_cq,
        )
    except:
        await MessageService.edith_managed_message(
                bot=message.bot,
                user_id=user_dbm.tg_id,
                text=blank.get_patient_confirmation_template(
                    survey_dbm=survey_dbm,
                    patient_dbm=patient_dbm,
                    doctor_id=user_dbm.tg_id,
                ),
                reply_markup=keyboard.get_patient_confirmation_keyboard(
                    patient_dbm=patient_dbm,
                    add_contact_button=False,
                ),
                state=state,
                previous_message_key="start_msg_id",
                message_id_storage_key="start_msg_id",
                new_state=ScheduleSurveyStates.waiting_confirm_selected_patient,
                message=message_from_cq,
        )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )  


@router.callback_query(F.data.startswith(DoctorAction.SELECT_PATIENT.value))
async def show_selected_patient_cq(
    call_back_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    patient_id = await MessageService.get_value_from_callback_data(
        callback_data=call_back_query.data,
        default_value=None,
    )
    await ScheduleSurveyService.save_selected_patient(
        state=state,
        patient_id=patient_id,
    )

    await proccess_show_selected_patient(
        message=call_back_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(ScheduleSurveyStates.waiting_confirm_selected_patient)
async def show_selected_patient(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    current_page = await ScheduleSurveyService.get_select_patient_current_page(state)

    await proccess_handle_patient_selection(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        page=current_page,
        from_cq=False,
    )



async def proccess_show_schedule_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
    from_cq: bool = True,
):
    message_from_cq = message if from_cq else None
    
    survey = await ScheduleSurveyService.get_survey(state)
    patient_dbm = await ScheduleSurveyService.get_selected_patient(state)

    await MessageService.edith_managed_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            text=blank.get_survey_planning_confirmation_blank(survey=survey),
            reply_markup=keyboard.get_schedule_survey_confirmation_keyboard(),
            state=state,
            previous_message_key="start_msg_id",
            message_id_storage_key="start_msg_id",
            new_state=ScheduleSurveyStates.waiting_confirm_schedule_survey,
            message=message_from_cq,
    )

    if not from_cq:
        await MessageService.remove_previous_message(
            bot=message.bot,
            user_id=user_dbm.tg_id,
            message_id=message.message_id,
        )  

@router.callback_query(F.data.startswith(DoctorAction.CONFIRM_SELECTED_PATIENT.value))
async def show_schedule_survey_cq(
    call_back_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    await proccess_show_schedule_survey(
        message=call_back_query.message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
    )

@router.message(ScheduleSurveyStates.waiting_confirm_schedule_survey)
async def show_schedule_survey(
    message: Message,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    await proccess_show_schedule_survey(
        message=message,
        state=state,
        keyboard=keyboard,
        blank=blank,
        user_dbm=user_dbm,
        from_cq=False,
    )

@router.callback_query(F.data.startswith(DoctorAction.CONFIRM_SCHEDULE_SURVEY.value))
async def confirm_schedule_survey_cq(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM],
):
    await callback_query.answer("Опрос успешно был запланирован!")

    await ScheduleSurveyService.schedule_suvey(
        state=state, 
        user_id=user_dbm.tg_id
    )

    await MessageService.edith_managed_message(
        bot=callback_query.bot,
        user_id=user_dbm.tg_id,
        text=blank.get_default_blank(user_dbm.full_name),
        reply_markup=keyboard.get_default_keyboard(),
        state=state,
        previous_message_key="start_msg_id",
        message_id_storage_key="start_msg_id",
        message=callback_query.message,
        new_state=None
    )



@router.callback_query(F.data.startswith(DoctorAction.GET_SURVEYS_STATICS.value))
async def handle_get_survey_statistics(
    callback_query: CallbackQuery,
    state: FSMContext,
    keyboard: type[DoctorKeyboard],
    blank: type[DoctorBlank],
    user_dbm: type[UserDBM]
):
    await callback_query.answer()
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Получаем все опросы
    surveys = await ScheduleSurveyService.get_all_surveys(user_dbm.tg_id)
    
    for survey in surveys:
        # Создаем лист для каждого опроса
        ws = wb.create_sheet(title=f"{survey.title[:25]}_{survey.id}")
        
        # Настройки форматирования
        ws.row_dimensions[1].height = 25
        alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
        
        # Получаем вопросы для этого опроса
        question_dbms = await ScheduleSurveyService.get_questions_by_survey(
            survey_id=survey.id
        )
        
        # Формируем заголовки
        headers = [
            "Пользователь",
            "Запланированная дата",
            "Запланированное время",
        ]

        # Добавляем вопросы в порядке их ID
        question_dbms_sorted = sorted(question_dbms, key=lambda x: x.id)
        for question_dbm in question_dbms_sorted:
            headers.append(f"Вопрос с ID: {question_dbm.id}")
        
        ws.append(headers)
    
        # Форматирование заголовков
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            
            # Автоматическое выравнивание ширины столбцов
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = max(
                len(str(cell.value)) + 2,  # +2 для отступов
                10  # Минимальная ширина
            )

        response_for_all_time = await ScheduleSurveyService.get_survey_responses_for_all_time(survey.id)

        for row_idx, row in enumerate(response_for_all_time, start=2):
            user_id, curr_date, scheduled_time, answers = row
            
            # Создаем строку для Excel
            excel_row = [user_id, curr_date, scheduled_time]
            excel_row.extend(answers)
            
            # Добавляем строку в лист
            ws.append(excel_row)
            
            # Форматирование ячеек данных
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.alignment = alignment
                cell.border = border
                
                # Обновляем ширину столбца, если значение шире текущего
                column_letter = get_column_letter(col_idx)
                current_width = ws.column_dimensions[column_letter].width
                value_length = len(str(cell.value)) + 2  # +2 для отступов
                if value_length > current_width:
                    ws.column_dimensions[column_letter].width = min(value_length, 50)  # Макс 50 символов
    
    # Сохраняем и отправляем файл
    file_path = f'./survey_stats_{uuid.uuid4()}.xlsx'
    wb.save(file_path)
    
    try:
        stat_file = FSInputFile(file_path)
        await callback_query.message.answer_document(
            document=stat_file,
            caption=f"Статистика по {len(surveys)} опросам"
        )
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)